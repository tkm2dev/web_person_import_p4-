# -*- coding: utf-8 -*-
"""
build_p4.py — ล้างข้อมูล + แยกไฟล์รายจังหวัด จาก person_import_template_p4.xlsx

หลักการสำคัญ: ห้ามเปลี่ยนโครงสร้างคอลัมน์ของ template
  - หัวตารางของทุกชีตอ่านจากไฟล์ต้นฉบับโดยตรง (ไม่ hardcode) จึงตรงกัน 100%
  - ชื่อชีตเดิม / ชื่อประเภทข้อมูลเดิม คงไว้ทุกตัว
  - ชีต README และ Lists คัดลอกมาทั้งแผ่น
  - 3 ประเภทที่ยังไม่มีชีต ยืมโครงคอลัมน์จากชีตที่มีอยู่ (ไม่คิดคอลัมน์ใหม่)
แก้เฉพาะ "ค่าในเซลล์" ที่ผิดจนแยกจังหวัดไม่ได้ และตัดแถวเปล่าออก

ผลลัพธ์ใน build/
  P4_SOURCE.xlsx              ต้นทางที่ล้างแล้ว (อัปโหลดขึ้น Drive ให้ Apps Script อ่าน)
  provinces/<จังหวัด>.xlsx    ไฟล์ตั้งต้นรายจังหวัด (สำรอง เผื่ออัปโหลดเอง)
  data_quality_report.md      รายงานสิ่งที่แก้ / สิ่งที่ยังขาด

  python scripts/build_p4.py
"""
from __future__ import annotations

import datetime as dt
import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path

warnings.filterwarnings("ignore")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "person_import_template_p4.xlsx"
BUILD = ROOT / "build"
PROV_DIR = BUILD / "provinces"

SPARE_ROWS = 500

# ---------------------------------------------------------------- schema
# src   = ชีตต้นฉบับที่มีข้อมูล (None = ชีตใหม่ ยังไม่มีข้อมูล)
# donor = ชีตที่ยืมโครงคอลัมน์มาใช้ สำหรับชีตใหม่
# type  = ค่าที่ต้องอยู่ในคอลัมน์ "ประเภทข้อมูล*" (ต้องตรงกับ dropdown ในชีต Lists)
#
# ลำดับใน dict = ลำดับแท็บในไฟล์ผลลัพธ์ (คงลำดับเดิมของต้นฉบับ แล้วต่อท้าย 3 ชีตใหม่)
SHEETS: dict[str, dict] = {
    "กลุ่ม5สี11กลุ่ม":        {"src": "กลุ่ม5สี11กลุ่ม", "type": "กลุ่ม 5 สี 11 กลุ่ม"},
    "ถวายฎีกา":               {"src": "ถวายฎีกา",        "type": "บุคคลถวายฎีกา"},
    "ม112_มั่นคง":            {"src": "ม112_มั่นคง",     "type": "บุคคล ม.112/คดีความมั่นคง"},
    "จิตเวชรักษา":            {"src": "จิตเวชรักษา",     "type": "บุคคลจิตเวชมีประวัติการรักษา"},
    "เร่ร้อน":                {"src": "เร่ร้อน",         "type": "บุคคลเร่ร้อน"},
    "ร้องทุกข์_ดำรงธรรม":     {"src": None, "donor": "ถวายฎีกา",
                               "type": "ข้อมูลยื่นเรื่องราวร้องทุกข์ผ่านศูนย์ดำรงธรรม"},
    "ร้องทุกข์_หน่วยงานอื่น": {"src": None, "donor": "ถวายฎีกา",
                               "type": "ข้อมูลยื่นเรื่องราวร้องทุกข์ หน่วยงานอื่นๆ"},
    "เฝ้าระวัง_ทะเลาะวิวาท":  {"src": None, "donor": "กลุ่ม5สี11กลุ่ม",
                               "type": "กลุ่มบุคคลเฝ้าระวัง (ทะเลาะวิวาท)"},
}

COL_TYPE = "ประเภทข้อมูล*"
COL_PROVINCE = "จังหวัด*"

PROVINCES = [
    "กาฬสินธุ์", "ขอนแก่น", "นครพนม", "บึงกาฬ", "มหาสารคาม", "มุกดาหาร",
    "ร้อยเอ็ด", "เลย", "สกลนคร", "หนองคาย", "หนองบัวลำภู", "อุดรธานี",
]
PROVINCE_TYPO = {"อุดารธานี": "อุดรธานี"}

DATE_COLS = {"วันที่บันทึก*", "วันเกิด", "วันที่ติดตามถัดไป"}
TEXT_COLS = {"เลขบัตรประชาชน", "เลขเอกสารอื่น", "โทรศัพท์", "เลขที่อ้างอิงต้นทาง",
             "เลขคำร้อง/เลขฎีกา"}

# หัวคอลัมน์ในชีตข้อมูล -> หัวรายการในชีต Lists (ใช้ทำ dropdown)
DROPDOWN = {
    COL_TYPE: "ประเภทข้อมูล",
    COL_PROVINCE: "จังหวัด",
    "คำนำหน้า": "คำนำหน้า",
    "เพศ": "เพศ",
    "ช่องทางรับเรื่อง": "ช่องทางรับเรื่อง",
}

README_EXTRA = [
    ("", ""),
    ("— ฉบับแยกจังหวัด (ภ.4) —", ""),
    ("ขอบเขตไฟล์นี้", "ไฟล์นี้เป็นของจังหวัดเดียว ช่อง 'จังหวัด*' ถูกล็อกไว้แล้ว"),
    ("ข้อมูลตั้งต้น", "แถวที่มีอยู่คือข้อมูลเดิมของจังหวัดท่าน โปรดตรวจสอบ แก้ไข และกรอกช่องที่ยังว่างให้ครบ"),
    ("ช่องบังคับที่ว่าง", "ช่องที่มี * และยังว่าง จะขึ้นพื้นสีแดงอัตโนมัติในเวอร์ชัน Google Sheet"),
    ("การส่งข้อมูล", "ไม่ต้องส่งไฟล์กลับ ภ.4 ดึงข้อมูลอัตโนมัติทุก 5 นาที"),
    ("ห้ามแก้โครงตาราง", "ห้ามเพิ่ม/ลบ/สลับคอลัมน์ และห้ามเปลี่ยนชื่อแท็บ เพราะ ภ.4 ต้องรวมส่งส่วนกลางด้วย template นี้"),
]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_REQ_FILL = PatternFill("solid", fgColor="7B2D26")
HDR_FONT = Font(color="FFFFFF", bold=True)


# ---------------------------------------------------------------- helpers

def is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def clean_scalar(v):
    if isinstance(v, str):
        v = v.replace("\r\n", "\n").strip()
        return v or None
    return v


def to_iso_date(v, log: Counter):
    """คืน 'yyyy-mm-dd'; ปี >= 2400 ถือเป็น พ.ศ. แปลงเป็น ค.ศ. (README กำหนดให้ใช้ ค.ศ.)"""
    if is_blank(v):
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        y = v.year
        if y >= 2400:
            log["be_to_ce"] += 1
            y -= 543
        try:
            return dt.date(y, v.month, v.day).isoformat()
        except ValueError:
            log["bad_date"] += 1
            return str(v)
    s = str(v).strip()
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        y, mo, d = int(m[1]), int(m[2]), int(m[3])
        if y >= 2400:
            log["be_to_ce"] += 1
            y -= 543
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            log["bad_date"] += 1
            return s
    log["unparsed_date"] += 1
    return s


def to_text_id(v):
    if is_blank(v):
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def quote_inline(value: str) -> str:
    """inline dropdown ค่าเดียวสำหรับ openpyxl DataValidation"""
    return '"' + value.replace('"', "'") + '"'


# ---------------------------------------------------------------- read source

def read_source():
    """คืน (headers_by_tab, rows_by_tab, lists_grid, readme_grid, report)"""
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)

    # หัวตารางจริงจากต้นฉบับ
    headers: dict[str, list[str]] = {}
    for tab, cfg in SHEETS.items():
        origin = cfg["src"] or cfg["donor"]
        ws = wb[origin]
        headers[tab] = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # ชีต Lists และ README คัดลอกทั้งแผ่น
    lists_grid = grid_of(wb["Lists"])
    readme_grid = grid_of(wb["README"])

    rep = {
        "scanned": {}, "padding": {}, "kept": {},
        "typo_prov": Counter(), "type_forced": Counter(), "date_log": Counter(),
        "unknown_prov": Counter(), "no_prov": Counter(),
        "missing_required": defaultdict(Counter),
    }
    rows_by_tab: dict[str, list[dict]] = {t: [] for t in SHEETS}

    for tab, cfg in SHEETS.items():
        if cfg["src"] is None:
            rep["scanned"][tab] = rep["padding"][tab] = rep["kept"][tab] = 0
            continue

        ws = wb[cfg["src"]]
        hdr = headers[tab]
        type_i = hdr.index(COL_TYPE)
        prov_i = hdr.index(COL_PROVINCE)
        scanned = padding = 0

        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(hdr) + 1)]
            if all(is_blank(v) for v in vals):
                continue
            scanned += 1
            # แถวเปล่าที่ pre-fill เฉพาะคอลัมน์ประเภทข้อมูล/จังหวัด
            if all(is_blank(v) for i, v in enumerate(vals) if i not in (type_i, prov_i)):
                padding += 1
                continue

            rec = {h: clean_scalar(vals[i]) for i, h in enumerate(hdr)}

            if rec[COL_TYPE] != cfg["type"]:
                rep["type_forced"][f"{tab}: {rec[COL_TYPE]!r} -> {cfg['type']!r}"] += 1
                rec[COL_TYPE] = cfg["type"]

            p = rec[COL_PROVINCE]
            if isinstance(p, str) and p in PROVINCE_TYPO:
                rep["typo_prov"][f"{p} -> {PROVINCE_TYPO[p]}"] += 1
                p = PROVINCE_TYPO[p]
            if is_blank(p):
                rep["no_prov"][tab] += 1
                p = None
            elif p not in PROVINCES:
                rep["unknown_prov"][f"{tab}: {p}"] += 1
            rec[COL_PROVINCE] = p

            for h in hdr:
                if h in DATE_COLS:
                    rec[h] = to_iso_date(rec[h], rep["date_log"])
                elif h in TEXT_COLS:
                    rec[h] = to_text_id(rec[h])
                if h.endswith("*") and is_blank(rec[h]):
                    rep["missing_required"][tab][h] += 1

            rows_by_tab[tab].append(rec)

        rep["scanned"][tab] = scanned
        rep["padding"][tab] = padding
        rep["kept"][tab] = len(rows_by_tab[tab])

    return headers, rows_by_tab, lists_grid, readme_grid, rep


def grid_of(ws) -> list[list]:
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]


def lists_layout(lists_grid) -> dict[str, tuple[str, int]]:
    """หัวรายการใน Lists -> (คอลัมน์, จำนวนรายการ)"""
    out = {}
    if not lists_grid:
        return out
    ncol = max(len(r) for r in lists_grid)
    for c in range(ncol):
        head = lists_grid[0][c] if c < len(lists_grid[0]) else None
        if is_blank(head):
            continue
        n = sum(1 for r in range(1, len(lists_grid))
                if c < len(lists_grid[r]) and not is_blank(lists_grid[r][c]))
        out[str(head)] = (get_column_letter(c + 1), n)
    return out


# ---------------------------------------------------------------- write

def write_grid(wb, name, grid, widths=None, bold_first_row=False):
    ws = wb.create_sheet(name)
    for r, row in enumerate(grid, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
    if bold_first_row:
        for c in range(1, (len(grid[0]) if grid else 0) + 1):
            ws.cell(1, c).font = Font(bold=True)
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def write_data_tab(wb, tab, headers, rows, layout, province=None):
    cfg = SHEETS[tab]
    ws = wb.create_sheet(tab)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_REQ_FILL if str(h).endswith("*") else HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 18 if len(str(h)) < 16 else 26

    for r, rec in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(r, c, rec.get(h))
            if h in TEXT_COLS:
                cell.number_format = "@"

    # แถวสำรอง: pre-fill ประเภทข้อมูล (+ จังหวัด ถ้าเป็นไฟล์จังหวัด) เหมือน template เดิม
    ti = headers.index(COL_TYPE) + 1
    pi = headers.index(COL_PROVINCE) + 1
    first_spare = len(rows) + 2
    last = len(rows) + 1 + SPARE_ROWS
    for r in range(first_spare, last + 1):
        ws.cell(r, ti, cfg["type"])
        if province:
            ws.cell(r, pi, province)
        for c, h in enumerate(headers, start=1):
            if h in TEXT_COLS:
                ws.cell(r, c).number_format = "@"

    ws.freeze_panes = "A2"

    for c, h in enumerate(headers, start=1):
        col = get_column_letter(c)
        rng = f"{col}2:{col}{last}"
        if h == COL_TYPE:
            dv = DataValidation(type="list", formula1=quote_inline(cfg["type"]), allow_blank=True)
        elif h == COL_PROVINCE and province:
            dv = DataValidation(type="list", formula1=quote_inline(province), allow_blank=True)
        elif h in DROPDOWN and DROPDOWN[h] in layout:
            letter, n = layout[DROPDOWN[h]]
            dv = DataValidation(type="list",
                                formula1=f"=Lists!${letter}$2:${letter}${n + 1}",
                                allow_blank=True)
        elif h in DATE_COLS:
            dv = DataValidation(type="date", operator="greaterThan",
                                formula1="DATE(1900,1,1)", allow_blank=True)
            dv.error = "ใช้รูปแบบวันที่ ค.ศ. yyyy-mm-dd"
        else:
            continue
        ws.add_data_validation(dv)
        dv.add(rng)
    return ws


def build_workbook(path: Path, headers, rows_by_tab, lists_grid, readme_grid,
                   layout, province=None, leftovers=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    grid = [list(r) for r in readme_grid]
    if province:
        grid = grid + [list(x) for x in README_EXTRA]
    rm = write_grid(wb, "README", grid, widths=[44, 110])
    rm.cell(1, 1).font = Font(bold=True, size=13)
    for r in range(1, len(grid) + 1):
        rm.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

    for tab in SHEETS:
        write_data_tab(wb, tab, headers[tab], rows_by_tab.get(tab, []), layout, province)

    write_grid(wb, "Lists", lists_grid, widths=[28] * 20, bold_first_row=True)

    if leftovers:
        ws = wb.create_sheet("_ตกค้าง")
        ws.cell(1, 1, "แท็บต้นทาง").font = Font(bold=True)
        for r, (tab, rec) in enumerate(leftovers, start=2):
            hdr = headers[tab]
            if r == 2:
                for c, h in enumerate(hdr, start=2):
                    ws.cell(1, c, h).font = Font(bold=True)
            ws.cell(r, 1, tab)
            for c, h in enumerate(hdr, start=2):
                ws.cell(r, c, rec.get(h))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------- report

def write_report(rep, headers, by_prov, leftovers):
    L = ["# รายงานคุณภาพข้อมูล — person_import_template_p4.xlsx", "",
         f"สร้างโดย `scripts/build_p4.py` | ต้นฉบับ: `{SRC_XLSX.name}`", "",
         "> โครงสร้างคอลัมน์ของทุกชีตคงเดิม 100% (อ่านหัวตารางจากไฟล์ต้นฉบับ) "
         "แก้เฉพาะค่าในเซลล์ที่ผิดจนแยกจังหวัดไม่ได้ และตัดแถวเปล่าออก", ""]

    L += ["## 1. จำนวนแถว", "",
          "| ชีต | คอลัมน์ | แถวที่อ่าน | แถวเปล่า | แถวที่เก็บ |", "|---|---:|---:|---:|---:|"]
    for tab in SHEETS:
        L.append(f"| {tab} | {len(headers[tab])} | {rep['scanned'][tab]:,} | "
                 f"{rep['padding'][tab]:,} | {rep['kept'][tab]:,} |")
    L.append(f"| **รวม** | | **{sum(rep['scanned'].values()):,}** | "
             f"**{sum(rep['padding'].values()):,}** | **{sum(rep['kept'].values()):,}** |")
    L += ["", "แถวเปล่า = แถวที่ pre-fill เฉพาะคอลัมน์ ประเภทข้อมูล*/จังหวัด* "
              "ไม่มีข้อมูลบุคคลจริง จึงถูกแทนที่ด้วยแถวสำรองเปล่าท้ายตาราง", ""]

    L += ["## 2. ค่าในเซลล์ที่แก้ให้อัตโนมัติ", ""]
    if rep["typo_prov"]:
        L.append("**ชื่อจังหวัดสะกดผิด** (ถ้าไม่แก้ แถวเหล่านี้จะแยกเข้าจังหวัดไม่ได้)")
        L += [f"- `{k}` — {v:,} แถว" for k, v in rep["typo_prov"].items()] + [""]
    if rep["type_forced"]:
        L.append("**ประเภทข้อมูลไม่ตรงชีต (บังคับตามชีต)**")
        L += [f"- {k} — {v:,} แถว" for k, v in rep["type_forced"].items()] + [""]
    if rep["date_log"]:
        names = {"be_to_ce": "แปลง พ.ศ. -> ค.ศ. ตามที่ README กำหนด",
                 "bad_date": "วันที่ไม่ถูกต้อง (คงค่าเดิมไว้)",
                 "unparsed_date": "แปลงรูปแบบไม่ได้ (เก็บเป็นข้อความ)"}
        L.append("**วันที่**")
        L += [f"- {names.get(k, k)} — {v:,} ค่า" for k, v in rep["date_log"].items()] + [""]
    L += ["**เลขบัตร/เลขเอกสาร**", "- บังคับเป็นข้อความทุกช่อง เลขศูนย์นำหน้าจะไม่หาย", ""]

    L += ["## 3. ชีตที่เพิ่มเข้ามา (ตามภาพประเภทข้อมูล 8 ประเภท)", "",
          "ไม่มีการคิดคอลัมน์ใหม่ — ยืมโครงคอลัมน์จากชีตที่มีอยู่ในไฟล์ต้นฉบับ", "",
          "| ชีตใหม่ | ยืมโครงจาก | จำนวนคอลัมน์ |", "|---|---|---:|"]
    for tab, cfg in SHEETS.items():
        if cfg["src"] is None:
            L.append(f"| {tab} | {cfg['donor']} | {len(headers[tab])} |")
    L.append("")

    L += ["## 4. เรื่องที่ยังต้องให้เจ้าหน้าที่กรอกเพิ่ม", "",
          "คอลัมน์บังคับ (`*`) ที่ยังว่างหลังล้างข้อมูล — จะขึ้นพื้นสีแดงในไฟล์จังหวัด", "",
          "| ชีต | คอลัมน์บังคับที่ว่าง | จำนวนแถว | จากทั้งหมด |", "|---|---|---:|---:|"]
    for tab, cnt in rep["missing_required"].items():
        for col, n in sorted(cnt.items(), key=lambda kv: -kv[1]):
            L.append(f"| {tab} | {col} | {n:,} | {rep['kept'][tab]:,} |")
    L.append("")

    if rep["no_prov"] or rep["unknown_prov"] or leftovers:
        L += ["## 5. แถวที่แยกจังหวัดไม่ได้", ""]
        for tab, n in rep["no_prov"].items():
            L.append(f"- {tab}: จังหวัดว่าง {n:,} แถว")
        for k, n in rep["unknown_prov"].items():
            L.append(f"- {k}: ไม่อยู่ใน 12 จังหวัด ภ.4 — {n:,} แถว")
        L += ["", f"เก็บไว้ในชีต `_ตกค้าง` ของ `P4_SOURCE.xlsx` ({len(leftovers):,} แถว)", ""]

    L += ["## 6. ข้อมูลตั้งต้นรายจังหวัด", "",
          "| จังหวัด | " + " | ".join(SHEETS) + " | รวม |",
          "|---|" + "---:|" * (len(SHEETS) + 1)]
    for p in PROVINCES:
        cells = [f"{len(by_prov[p][t]):,}" for t in SHEETS]
        total = sum(len(by_prov[p][t]) for t in SHEETS)
        L.append(f"| {p} | " + " | ".join(cells) + f" | **{total:,}** |")
    tot = [f"**{sum(len(by_prov[p][t]) for p in PROVINCES):,}**" for t in SHEETS]
    grand = sum(sum(len(by_prov[p][t]) for t in SHEETS) for p in PROVINCES)
    L.append("| **รวม** | " + " | ".join(tot) + f" | **{grand:,}** |")

    L += ["", "## 7. เรื่องที่ยังต้องยืนยันกับส่วนกลาง", "",
          "- ชีต `Lists` ระบุประเภท `บุคคลถวายฎีกา` แต่ README เขียนว่า "
          "`บุคคลถวายฎีกาผ่านศูนย์ดำรงธรรม` — สคริปต์ใช้ค่าตาม `Lists` (ตรงกับภาพประเภทข้อมูล)",
          "- ชีต `เร่ร้อน` / ประเภท `บุคคลเร่ร้อน` สะกดต่างจากภาพประเภทข้อมูลที่เขียนว่า "
          "`บุคคลเร่ร่อน` — คงตามไฟล์ต้นฉบับไว้ ไม่แก้",
          "- README อ้างถึงชีต `นำเข้ารวม` ที่ไม่มีอยู่จริงในไฟล์",
          "- ชีต `Lists` ไม่มีรายการค่าสำหรับ `ระดับความเร่งด่วน` จึงยังเป็นช่องพิมพ์อิสระ",
          "- ชีต `จิตเวชรักษา` ใช้หัวคอลัมน์ `ชื่อ` (ไม่มี `*`) ต่างจากชีตอื่นที่ใช้ `ชื่อ*` — คงไว้ตามต้นฉบับ",
          ""]

    (BUILD / "data_quality_report.md").write_text("\n".join(L), encoding="utf-8")


# ---------------------------------------------------------------- main

def main():
    print(f"อ่าน {SRC_XLSX}")
    headers, rows_by_tab, lists_grid, readme_grid, rep = read_source()
    layout = lists_layout(lists_grid)
    print("รายการใน Lists:", {k: v[1] for k, v in layout.items()})

    by_prov = {p: {t: [] for t in SHEETS} for p in PROVINCES}
    leftovers: list[tuple[str, dict]] = []
    for tab, rows in rows_by_tab.items():
        for rec in rows:
            p = rec[COL_PROVINCE]
            if p in by_prov:
                by_prov[p][tab].append(rec)
            else:
                leftovers.append((tab, rec))

    BUILD.mkdir(exist_ok=True)
    print("เขียน build/P4_SOURCE.xlsx")
    build_workbook(BUILD / "P4_SOURCE.xlsx", headers, rows_by_tab, lists_grid,
                   readme_grid, layout, province=None, leftovers=leftovers)

    PROV_DIR.mkdir(parents=True, exist_ok=True)
    for p in PROVINCES:
        n = sum(len(by_prov[p][t]) for t in SHEETS)
        print(f"เขียน build/provinces/{p}.xlsx ({n:,} แถว)")
        build_workbook(PROV_DIR / f"{p}.xlsx", headers, by_prov[p], lists_grid,
                       readme_grid, layout, province=p)

    write_report(rep, headers, by_prov, leftovers)
    print("เขียน build/data_quality_report.md")
    print(f"\nเสร็จ — ข้อมูลจริง {sum(rep['kept'].values()):,} แถว, ตกค้าง {len(leftovers):,} แถว")


if __name__ == "__main__":
    main()
